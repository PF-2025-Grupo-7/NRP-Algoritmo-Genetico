import json
import traceback
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
from django.urls import reverse_lazy
from django.utils import timezone
from django.contrib import messages
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST, require_GET
from django.core.exceptions import ValidationError
from django.views.generic import (
    ListView, CreateView, UpdateView, DeleteView, 
    DetailView, TemplateView, FormView
)
from django.contrib.auth.decorators import user_passes_test
from django.core.exceptions import PermissionDenied

# --- MODELOS ---
from .models import (
    Empleado, Cronograma, TipoTurno, NoDisponibilidad, 
    Preferencia, SecuenciaProhibida, Asignacion,
    PlantillaDemanda, ReglaDemandaSemanal, ExcepcionDemanda,
    TrabajoPlanificacion, ConfiguracionAlgoritmo
)

# --- FORMS ---
from .forms import (
    EmpleadoForm, ConfiguracionTurnosForm, NoDisponibilidadForm, 
    PreferenciaForm, SecuenciaProhibidaForm,
    PlantillaDemandaForm, ReglaDemandaSemanalForm, ExcepcionDemandaForm,
    ConfiguracionSimpleForm, ConfiguracionAvanzadaForm
)

# --- FILTROS ---
from .filters import (
    EmpleadoFilter, CronogramaFilter, NoDisponibilidadFilter, 
    PreferenciaFilter, TipoTurnoFilter, SecuenciaProhibidaFilter
)

# --- SERVICIOS (Lógica de Negocio) ---
from .services import (
    iniciar_proceso_optimizacion, # <--- Nueva función orquestadora
    consultar_resultado_ag, 
    guardar_solucion_db,
    invocar_api_planificacion,
    construir_matriz_cronograma   # <--- Nueva función de presentación
)

class SuperUserRequiredMixin(UserPassesTestMixin):
    raise_exception = True  # <--- ESTA LÍNEA ES LA CLAVE MÁGICA
    def test_func(self): 
        return self.request.user.is_superuser
    
# ==============================================================================
# VISTAS GENERALES Y DE GESTIÓN
# ==============================================================================

def landing(request):
    """Pantalla de presentación: redirige a dashboard si está logueado."""
    if request.user.is_authenticated:
        return redirect('dashboard')
    return render(request, 'rostering/landing.html')


def dashboard(request):
    """Vista principal: KPIs, accesos rápidos y estado actual del sistema."""
    total_empleados = Empleado.objects.count()
    total_ausencias = NoDisponibilidad.objects.count()
    total_preferencias = Preferencia.objects.count()
    recientes = Cronograma.objects.all().order_by('-fecha_creacion')[:5]

    hoy = timezone.now().date()
    if hoy.month == 12:
        prox_mes, prox_anio = 1, hoy.year + 1
    else:
        prox_mes, prox_anio = hoy.month + 1, hoy.year
    
    existe_proximo = Cronograma.objects.filter(fecha_inicio__month=prox_mes, fecha_inicio__year=prox_anio).exists()
    nombre_mes_objetivo = f"{prox_mes:02d}/{prox_anio}"

    context = {
        'total_empleados': total_empleados,
        'total_ausencias': total_ausencias,
        'total_preferencias': total_preferencias,
        'recientes': recientes,
        'mes_objetivo': nombre_mes_objetivo,
    }
    return render(request, 'rostering/dashboard.html', context)


def registrar_usuario(request):
    """Permite el registro de nuevos usuarios en el sistema."""
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, "¡Registro exitoso! Bienvenido.")
            return redirect('dashboard')
    else:
        form = UserCreationForm()
    return render(request, 'registration/register.html', {'form': form})


# ==============================================================================
# MOTOR DE PLANIFICACIÓN (API ORCHESTRATION)
# ==============================================================================

@login_required
def pagina_generador(request):
    """Renderiza la pantalla para configurar y lanzar una nueva planificación."""
    return render(request, 'rostering/generador.html')


@csrf_exempt 
@login_required
@require_POST
def iniciar_planificacion(request):
    try:
        data = json.loads(request.body)
        job_id = iniciar_proceso_optimizacion(data)
        return JsonResponse({'status': 'started', 'job_id': job_id})

    except ValidationError as e:
        # Si el error trae parámetros (nuestros datos de dotación), los enviamos
        if hasattr(e, 'params') and e.params:
            return JsonResponse({
                'error': e.message,
                'tipo_error': 'FALTA_DOTACION', # Flag para que el JS sepa qué modal abrir
                'detalles': e.params
            }, status=422) # 422: Entidad no procesable (Validación de negocio fallida)
        
        # Error genérico de validación
        return JsonResponse({'error': str(e.message)}, status=400)
        
    except ValueError as e:
        return JsonResponse({'error': str(e)}, status=400)
    except Exception as e:
        print(e)
        return JsonResponse({'error': "Error interno del servidor"}, status=500)


@csrf_exempt
@require_GET
def verificar_estado_planificacion(request, job_id):
    """
    Polling: El frontend consulta esto periódicamente.
    """
    try:
        try:
            trabajo = TrabajoPlanificacion.objects.get(job_id=job_id)
        except TrabajoPlanificacion.DoesNotExist:
            return JsonResponse({'error': 'Job ID no encontrado o expirado.'}, status=404)

        resultado = consultar_resultado_ag(job_id)
        
        if not resultado:
            return JsonResponse({'status': 'error', 'mensaje': 'Error de conexión'}, status=503)

        if resultado.get('status') == 'error':
             return JsonResponse({'status': 'failed', 'error': resultado.get('error')})

        # Si terminó, persistimos
        if 'fitness' in resultado or resultado.get('status') == 'completed':
            try:
                cronograma = guardar_solucion_db(
                    fecha_inicio=trabajo.fecha_inicio, 
                    fecha_fin=trabajo.fecha_fin, 
                    especialidad=trabajo.especialidad, 
                    payload_original=trabajo.payload_original, 
                    resultado=resultado,
                    plantilla_demanda=trabajo.plantilla_demanda
                )
                trabajo.delete() # Limpieza

                return JsonResponse({
                    'status': 'completed',
                    'cronograma_id': cronograma.id,
                    'fitness': resultado.get('fitness'),
                    'mensaje': 'Planificación guardada con éxito.'
                })
            except Exception as e:
                # Esto imprimirá el error real en tu consola de Docker
                print("🔴 ERROR CRÍTICO EN POLLING/GUARDADO:")
                print(traceback.format_exc()) 
                return JsonResponse({'error': f"Error interno: {str(e)}"}, status=500)

        return JsonResponse({'status': 'running'})

    except Exception as e:
        print(f"Error polling: {e}")
        return JsonResponse({'error': str(e)}, status=500)


def api_get_plantillas(request):
    especialidad = request.GET.get('especialidad')
    if not especialidad: return JsonResponse({'plantillas': []})
    plantillas = PlantillaDemanda.objects.filter(especialidad=especialidad).values('id', 'nombre')
    return JsonResponse({'plantillas': list(plantillas)})


# ==============================================================================
# VISUALIZACIÓN DE CRONOGRAMAS
# ==============================================================================

@login_required
def ver_cronograma(request, cronograma_id):
    """
    Vista Detallada (Matriz): Ahora solo pide los datos procesados al servicio.
    INCLUYE BLOQUEO DE SEGURIDAD PARA CRONOGRAMAS FALLIDOS.
    """
    cronograma = get_object_or_404(Cronograma, pk=cronograma_id)
    
    # --- BLOQUEO DE SEGURIDAD ---
    if cronograma.estado == 'FALLIDO':
        messages.warning(request, "Acceso denegado: El cronograma no es viable. Se redirigió al análisis de errores.")
        return redirect('cronograma_analisis', pk=cronograma.id)
    # ----------------------------
    
    # Delegamos la construcción de la matriz al servicio
    context_data = construir_matriz_cronograma(cronograma)
    
    return render(request, 'rostering/cronograma_detail.html', {
        'cronograma': cronograma,
        'filter_form': None, # Si usabas filtros en el detalle, sino borralo
        **context_data 
    })

def ver_cronograma_diario(request, cronograma_id):
    """Vista Diaria: Muestra quién trabaja en qué turno para cada día."""
    cronograma = get_object_or_404(Cronograma, pk=cronograma_id)
    
    # --- BLOQUEO DE SEGURIDAD ---
    if cronograma.estado == 'FALLIDO':
        messages.warning(request, "Acceso denegado: El cronograma no es viable. Se redirigió al análisis de errores.")
        return redirect('cronograma_analisis', pk=cronograma.id)
    # ----------------------------
    
    asignaciones = Asignacion.objects.filter(cronograma=cronograma).select_related(
        'empleado', 'tipo_turno'
    ).order_by('fecha', 'tipo_turno__hora_inicio')

    agenda = {}
    for asig in asignaciones:
        if asig.fecha not in agenda: agenda[asig.fecha] = {}
        t_nom = asig.tipo_turno.nombre
        if t_nom not in agenda[asig.fecha]: agenda[asig.fecha][t_nom] = []
        agenda[asig.fecha][t_nom].append(asig.empleado)

    return render(request, 'rostering/cronograma_diario.html', {'cronograma': cronograma, 'agenda': agenda})

class CronogramaAnalisisView(LoginRequiredMixin, DetailView):
    model = Cronograma
    template_name = 'rostering/cronograma_analisis.html'
    context_object_name = 'cronograma'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        reporte = self.object.reporte_analisis or {}
        
        context['metricas'] = reporte.get('metricas', {})
        
        # Ensure we pass the dictionaries even if empty, so the template loop works
        violaciones = reporte.get('violaciones_blandas', {})
        context['violaciones_blandas'] = {
            'preferencia_libre_incumplida': violaciones.get('preferencia_libre_incumplida', []),
            'preferencia_turno_incumplida': violaciones.get('preferencia_turno_incumplida', [])
        }
        
        context['violaciones_duras'] = reporte.get('violaciones_duras', {})
        context['equidad'] = reporte.get('datos_equidad', {})
        return context

@login_required
@require_POST
def publicar_cronograma(request, pk):
    cronograma = get_object_or_404(Cronograma, pk=pk)
    
    # --- BLOQUEO ---
    if cronograma.estado == 'FALLIDO':
        messages.error(request, "Error crítico: No se puede publicar un cronograma marcado como NO VIABLE.")
        return redirect('cronograma_analisis', pk=cronograma.id)
    # ---------------

    if cronograma.estado == Cronograma.Estado.PUBLICADO:
        messages.warning(request, "Este cronograma ya está publicado.")
    else:
        cronograma.estado = Cronograma.Estado.PUBLICADO
        cronograma.save()
        messages.success(request, f"¡Planificación {cronograma.fecha_inicio} PUBLICADA!")
    return redirect('ver_cronograma', cronograma_id=pk)

# ==============================================================================
# ABM (CRUDs) - SIN CAMBIOS IMPORTANTES (Ya usan CBVs estándar)
# ==============================================================================

class EmpleadoListView(LoginRequiredMixin, ListView):
    model = Empleado
    template_name = 'rostering/empleado_list.html'
    context_object_name = 'empleados'
    paginate_by = 15 

    def get_queryset(self):
        qs = super().get_queryset().order_by('-activo', '-id')
        self.filterset = EmpleadoFilter(self.request.GET, queryset=qs)
        qs = self.filterset.qs
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['filter_form'] = self.filterset.form
        return ctx

class EmpleadoCreateView(LoginRequiredMixin, CreateView):
    model = Empleado
    form_class = EmpleadoForm
    template_name = 'rostering/empleado_form.html'
    success_url = reverse_lazy('empleado_list')
    extra_context = {'titulo': 'Nuevo Empleado'}
    
    def get_initial(self):
        initial = super().get_initial()
        initial['activo'] = True
        return initial

class EmpleadoUpdateView(LoginRequiredMixin, UpdateView):
    model = Empleado
    form_class = EmpleadoForm
    template_name = 'rostering/empleado_form.html'
    success_url = reverse_lazy('empleado_list')
    extra_context = {'titulo': 'Editar Empleado'}

class EmpleadoDeleteView(SuperUserRequiredMixin, DeleteView):
    model = Empleado
    template_name = 'rostering/empleado_confirm_delete.html'
    success_url = reverse_lazy('empleado_list')

# --- Cronogramas ---
class CronogramaListView(LoginRequiredMixin, ListView):
    model = Cronograma
    template_name = 'rostering/cronograma_list.html'
    context_object_name = 'cronogramas'
    ordering = ['-fecha_inicio', '-fecha_creacion'] 
    paginate_by = 10

    def get_queryset(self):
        qs = super().get_queryset()
        self.filterset = CronogramaFilter(self.request.GET, queryset=qs)
        return self.filterset.qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['filter_form'] = self.filterset.form
        return ctx

class CronogramaDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = Cronograma
    template_name = 'rostering/cronograma_confirm_delete.html'
    success_url = reverse_lazy('cronograma_list')

    def test_func(self):
        """
        Regla de Seguridad:
        1. Si es Superusuario -> Puede borrar cualquier cosa (Publicado o Borrador).
        2. Si es Mortal -> Solo puede borrar si el estado es BORRADOR.
        """
        cronograma = self.get_object()
        
        # Si es Admin, pase libre
        if self.request.user.is_superuser:
            return True
            
        # Si no es Admin, solo pasa si es Borrador
        return cronograma.estado == Cronograma.Estado.BORRADOR

# En views.py

from datetime import datetime, timedelta
from .models import ConfiguracionTurnos, TipoTurno
from .forms import ConfiguracionTurnosForm

class ConfiguracionTurnosListView(SuperUserRequiredMixin, ListView):
    """Muestra tarjetas por especialidad para entrar a configurar."""
    model = ConfiguracionTurnos
    template_name = 'rostering/config_turnos_list.html' # Nuevo template
    
    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        # Identificar qué especialidades ya tienen config
        configs = {c.especialidad: c for c in self.object_list}
        
        # Generar lista maestra de especialidades
        lista = []
        for esp_code, esp_label in Empleado.TipoEspecialidad.choices:
            lista.append({
                'code': esp_code,
                'label': esp_label,
                'config': configs.get(esp_code)
            })
        ctx['especialidades'] = lista
        return ctx

@login_required
def config_turnos_edit(request, especialidad):
    if not request.user.is_superuser:
        raise PermissionDenied("Solo los administradores pueden modificar la estructura de turnos.")
    
    instance = ConfiguracionTurnos.objects.filter(especialidad=especialidad).first()
    es_edicion = (instance is not None)
    
    if request.method == 'POST':
        form = ConfiguracionTurnosForm(request.POST, instance=instance)
        
        if form.is_valid():
            with transaction.atomic():
                config = form.save(commit=False)
                config.especialidad = especialidad
                
                nombres = {
                    "t1": {"n": form.cleaned_data['nombre_t1'], "a": form.cleaned_data['abrev_t1'], "noc": form.cleaned_data['nocturno_t1']},
                    "t2": {"n": form.cleaned_data['nombre_t2'], "a": form.cleaned_data['abrev_t2'], "noc": form.cleaned_data['nocturno_t2']},
                }
                if config.esquema == '3x8':
                    nombres["t3"] = {"n": form.cleaned_data['nombre_t3'], "a": form.cleaned_data['abrev_t3'], "noc": form.cleaned_data['nocturno_t3']}
                
                config.nombres_turnos = nombres
                config.save()

                # --- CÁLCULO DE HORARIOS ---
                hora_base = config.hora_inicio_base
                d = datetime(2000, 1, 1, hora_base.hour, hora_base.minute)
                
                nuevos_datos = []
                if config.esquema == '2x12':
                    nuevos_datos = [
                        {'key': 't1', 'inicio': d.time(), 'fin': (d + timedelta(hours=12)).time()},
                        {'key': 't2', 'inicio': (d + timedelta(hours=12)).time(), 'fin': (d + timedelta(hours=24)).time()}
                    ]
                elif config.esquema == '3x8':
                    for i in range(3):
                        nuevos_datos.append({
                            'key': f"t{i+1}", 
                            'inicio': (d + timedelta(hours=i*8)).time(), 
                            'fin': (d + timedelta(hours=(i+1)*8)).time()
                        })

                # --- CREACIÓN O ACTUALIZACIÓN ---
                if not es_edicion:
                    # CASO 1: PRIMERA VEZ (Crear objetos TipoTurno)
                    print(f"✨ CREACIÓN INICIAL: Generando turnos para {especialidad}")
                    for datos in nuevos_datos:
                        meta = nombres[datos['key']]
                        es_noc_calc = datos['fin'] < datos['inicio']
                        es_noc_final = True if es_noc_calc else meta['noc']

                        # CORRECCIÓN: Eliminamos activo=True
                        TipoTurno.objects.create(
                            nombre=meta['n'], abreviatura=meta['a'], especialidad=especialidad,
                            hora_inicio=datos['inicio'], hora_fin=datos['fin'], 
                            es_nocturno=es_noc_final
                        )
                    
                    regenerar_secuencias(especialidad)

                else:
                    # CASO 2: EDICIÓN (Actualizar existentes)
                    print(f"🔄 ACTUALIZACIÓN: Modificando detalles para {especialidad}")
                    
                    # CORRECCIÓN: Eliminamos activo=True del filtro
                    turnos_existentes = list(TipoTurno.objects.filter(especialidad=especialidad).order_by('hora_inicio'))
                    
                    if len(turnos_existentes) == len(nuevos_datos):
                        for idx, turno_obj in enumerate(turnos_existentes):
                            datos = nuevos_datos[idx]
                            meta = nombres[datos['key']]
                            
                            turno_obj.nombre = meta['n']
                            turno_obj.abreviatura = meta['a']
                            
                            es_noc_calc = datos['fin'] < datos['inicio']
                            turno_obj.es_nocturno = True if es_noc_calc else meta['noc']
                            
                            turno_obj.hora_inicio = datos['inicio']
                            turno_obj.hora_fin = datos['fin']
                            turno_obj.save()
                        
                        regenerar_secuencias(especialidad)
                    else:
                        print("⚠️ ERROR DE INTEGRIDAD: Cantidad de turnos en BD no coincide con esquema.")

            return redirect('tipoturno_list')
            
    else:
        initial_data = {}
        if instance and instance.nombres_turnos:
            nt = instance.nombres_turnos
            if 't1' in nt: initial_data.update({'nombre_t1': nt['t1']['n'], 'abrev_t1': nt['t1']['a'], 'nocturno_t1': nt['t1'].get('noc', False)})
            if 't2' in nt: initial_data.update({'nombre_t2': nt['t2']['n'], 'abrev_t2': nt['t2']['a'], 'nocturno_t2': nt['t2'].get('noc', False)})
            if 't3' in nt: initial_data.update({'nombre_t3': nt['t3']['n'], 'abrev_t3': nt['t3']['a'], 'nocturno_t3': nt['t3'].get('noc', False)})
        
        form = ConfiguracionTurnosForm(instance=instance, initial=initial_data)

    # CORRECCIÓN: Definimos la lista de horas aquí
    horas = [f"{h:02d}:{m:02d}" for h in range(0,24) for m in (0,30)]

    context = {
        'form': form,
        'especialidad_label': dict(Empleado.TipoEspecialidad.choices).get(especialidad),
        'es_edicion': es_edicion,
        'horas': horas 
    }
    return render(request, 'rostering/config_turnos_form.html', context)

def regenerar_secuencias(especialidad):
    """Helper para regenerar secuencias prohibidas."""
    SecuenciaProhibida.objects.filter(especialidad=especialidad).delete()
    
    # CORRECCIÓN: Eliminamos activo=True
    turnos_ordenados = list(TipoTurno.objects.filter(especialidad=especialidad).order_by('hora_inicio'))
    
    secuencias = []

    if len(turnos_ordenados) == 2: # 2x12
        secuencias.append(SecuenciaProhibida(especialidad=especialidad, turno_previo=turnos_ordenados[1], turno_siguiente=turnos_ordenados[0]))
    elif len(turnos_ordenados) == 3: # 3x8
        secuencias.append(SecuenciaProhibida(especialidad=especialidad, turno_previo=turnos_ordenados[2], turno_siguiente=turnos_ordenados[0]))
        secuencias.append(SecuenciaProhibida(especialidad=especialidad, turno_previo=turnos_ordenados[2], turno_siguiente=turnos_ordenados[1]))
        secuencias.append(SecuenciaProhibida(especialidad=especialidad, turno_previo=turnos_ordenados[1], turno_siguiente=turnos_ordenados[0]))

    if secuencias:
        SecuenciaProhibida.objects.bulk_create(secuencias)
    

# --- Ausencias ---
class NoDisponibilidadListView(LoginRequiredMixin, ListView):
    model = NoDisponibilidad
    template_name = 'rostering/nodisponibilidad_list.html'
    context_object_name = 'ausencias'
    ordering = ['-fecha_inicio']
    paginate_by = 10

    def get_queryset(self):
        qs = super().get_queryset().select_related('empleado')
        self.filterset = NoDisponibilidadFilter(self.request.GET, queryset=qs)
        return self.filterset.qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['filter_form'] = self.filterset.form
        return ctx

class NoDisponibilidadCreateView(LoginRequiredMixin, CreateView):
    model = NoDisponibilidad
    form_class = NoDisponibilidadForm
    template_name = 'rostering/nodisponibilidad_form.html'
    success_url = reverse_lazy('nodisponibilidad_list')
    extra_context = {'titulo': 'Registrar Ausencia'}

class NoDisponibilidadUpdateView(LoginRequiredMixin, UpdateView):
    model = NoDisponibilidad
    form_class = NoDisponibilidadForm
    template_name = 'rostering/nodisponibilidad_form.html'
    success_url = reverse_lazy('nodisponibilidad_list')
    extra_context = {'titulo': 'Editar Ausencia'}

class NoDisponibilidadDeleteView(LoginRequiredMixin, DeleteView):
    model = NoDisponibilidad
    template_name = 'rostering/confirm_delete_generic.html'
    success_url = reverse_lazy('nodisponibilidad_list')

# --- Preferencias ---
class PreferenciaListView(LoginRequiredMixin, ListView):
    model = Preferencia
    template_name = 'rostering/preferencia_list.html'
    context_object_name = 'preferencias'
    ordering = ['-fecha']
    paginate_by = 10

    def get_queryset(self):
        qs = super().get_queryset().select_related('empleado')
        self.filterset = PreferenciaFilter(self.request.GET, queryset=qs)
        return self.filterset.qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['filter_form'] = self.filterset.form
        return ctx

class PreferenciaCreateView(LoginRequiredMixin, CreateView):
    model = Preferencia
    form_class = PreferenciaForm
    template_name = 'rostering/preferencia_form.html'
    success_url = reverse_lazy('preferencia_list')
    extra_context = {'titulo': 'Nueva Preferencia'}

class PreferenciaUpdateView(LoginRequiredMixin, UpdateView):
    model = Preferencia
    form_class = PreferenciaForm
    template_name = 'rostering/preferencia_form.html'
    success_url = reverse_lazy('preferencia_list')
    extra_context = {'titulo': 'Editar Preferencia'}

class PreferenciaDeleteView(LoginRequiredMixin, DeleteView):
    model = Preferencia
    template_name = 'rostering/confirm_delete_generic.html'
    success_url = reverse_lazy('preferencia_list')


# --- Plantillas ---
class PlantillaListView(LoginRequiredMixin, ListView):
    model = PlantillaDemanda
    template_name = 'rostering/plantilla_list.html'
    context_object_name = 'plantillas'

class PlantillaCreateView(LoginRequiredMixin, CreateView):
    model = PlantillaDemanda
    form_class = PlantillaDemandaForm
    template_name = 'rostering/plantilla_form.html'
    success_url = reverse_lazy('plantilla_list')
    extra_context = {'titulo': 'Nueva Plantilla'}

class PlantillaDetailView(LoginRequiredMixin, DetailView):
    model = PlantillaDemanda
    template_name = 'rostering/plantilla_detail.html'
    context_object_name = 'plantilla'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['reglas'] = self.object.reglas.all().order_by('turno__hora_inicio')
        ctx['excepciones'] = self.object.excepciones.all().order_by('fecha')
        # Pasar todos los turnos de la especialidad para el dropdown
        ctx['turnos'] = TipoTurno.objects.filter(especialidad=self.object.especialidad).order_by('hora_inicio')
        return ctx

# Asegurate de importar el form nuevo arriba
from .forms import PlantillaDemandaUpdateForm 

class PlantillaUpdateView(LoginRequiredMixin, UpdateView):
    model = PlantillaDemanda
    form_class = PlantillaDemandaUpdateForm
    template_name = 'rostering/plantilla_form.html' # Reutilizamos el template de form
    
    def get_success_url(self):
        # Al guardar, volvemos al detalle de la plantilla
        return reverse_lazy('plantilla_detail', kwargs={'pk': self.object.pk})
        
    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['titulo'] = f"Editar: {self.object.nombre}"
        return ctx

class PlantillaDeleteView(LoginRequiredMixin, DeleteView):
    model = PlantillaDemanda
    template_name = 'rostering/confirm_delete_generic.html'
    success_url = reverse_lazy('plantilla_list')

class ReglaCreateView(LoginRequiredMixin, CreateView):
    model = ReglaDemandaSemanal
    form_class = ReglaDemandaSemanalForm
    template_name = 'rostering/regla_form.html'

    def get_form_kwargs(self):
        kw = super().get_form_kwargs()
        kw['plantilla_id'] = self.kwargs['plantilla_id']
        return kw

    def form_valid(self, form):
        form.instance.plantilla = PlantillaDemanda.objects.get(pk=self.kwargs['plantilla_id'])
        
        # Resolver conflictos: quitar días de otras reglas con el mismo turno
        self._resolver_conflictos_dias(form)
        
        return super().form_valid(form)
    
    def _resolver_conflictos_dias(self, form):
        """Remueve los días seleccionados de otras reglas que tengan el mismo turno."""
        dias_nuevos = form.cleaned_data['dias']
        turno = form.instance.turno
        plantilla = form.instance.plantilla
        
        # Buscar otras reglas con el mismo turno en la misma plantilla
        otras_reglas = ReglaDemandaSemanal.objects.filter(
            plantilla=plantilla,
            turno=turno
        ).exclude(pk=form.instance.pk if form.instance.pk else None)
        
        for regla in otras_reglas:
            dias_actuales = regla.dias or []
            # Quitar los días que están en la nueva regla
            dias_actualizados = [d for d in dias_actuales if d not in dias_nuevos]
            
            if dias_actualizados != dias_actuales:
                if dias_actualizados:  # Si quedan días, actualizar
                    regla.dias = dias_actualizados
                    regla.save()
                else:  # Si no quedan días, eliminar la regla
                    regla.delete()

    def get_success_url(self):
        return reverse_lazy('plantilla_detail', kwargs={'pk': self.kwargs['plantilla_id']})
    
    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        p = PlantillaDemanda.objects.get(pk=self.kwargs['plantilla_id'])
        ctx['titulo'] = f"Agregar Regla a {p.nombre}"
        return ctx

class ReglaUpdateView(LoginRequiredMixin, UpdateView):
    model = ReglaDemandaSemanal
    form_class = ReglaDemandaSemanalForm
    template_name = 'rostering/regla_form.html'

    def get_form_kwargs(self):
        kw = super().get_form_kwargs()
        kw['plantilla_id'] = self.object.plantilla.id
        return kw

    def form_valid(self, form):
        # Resolver conflictos: quitar días de otras reglas con el mismo turno
        self._resolver_conflictos_dias(form)
        return super().form_valid(form)
    
    def _resolver_conflictos_dias(self, form):
        """Remueve los días seleccionados de otras reglas que tengan el mismo turno."""
        dias_nuevos = form.cleaned_data['dias']
        turno = form.instance.turno
        plantilla = form.instance.plantilla
        
        # Buscar otras reglas con el mismo turno en la misma plantilla
        otras_reglas = ReglaDemandaSemanal.objects.filter(
            plantilla=plantilla,
            turno=turno
        ).exclude(pk=form.instance.pk)
        
        for regla in otras_reglas:
            dias_actuales = regla.dias or []
            # Quitar los días que están en la nueva regla
            dias_actualizados = [d for d in dias_actuales if d not in dias_nuevos]
            
            if dias_actualizados != dias_actuales:
                if dias_actualizados:  # Si quedan días, actualizar
                    regla.dias = dias_actualizados
                    regla.save()
                else:  # Si no quedan días, eliminar la regla
                    regla.delete()

    def get_success_url(self):
        return reverse_lazy('plantilla_detail', kwargs={'pk': self.object.plantilla.id})
    
    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['titulo'] = f"Editar Regla: {self.object.turno.nombre}"
        return ctx

class ReglaDeleteView(LoginRequiredMixin, DeleteView):
    model = ReglaDemandaSemanal
    template_name = 'rostering/confirm_delete_generic.html'
    def get_success_url(self):
        return reverse_lazy('plantilla_detail', kwargs={'pk': self.object.plantilla.id})

class ExcepcionCreateView(LoginRequiredMixin, CreateView):
    model = ExcepcionDemanda
    form_class = ExcepcionDemandaForm
    template_name = 'rostering/regla_form.html'

    def get_form_kwargs(self):
        kw = super().get_form_kwargs()
        kw['plantilla_id'] = self.kwargs['plantilla_id']
        return kw

    def form_valid(self, form):
        form.instance.plantilla = PlantillaDemanda.objects.get(pk=self.kwargs['plantilla_id'])
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('plantilla_detail', kwargs={'pk': self.kwargs['plantilla_id']})

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        p = PlantillaDemanda.objects.get(pk=self.kwargs['plantilla_id'])
        ctx['titulo'] = f"Agregar Excepción a {p.nombre}"
        return ctx

class ExcepcionDeleteView(LoginRequiredMixin, DeleteView):
    model = ExcepcionDemanda
    template_name = 'rostering/confirm_delete_generic.html'
    def get_success_url(self):
        return reverse_lazy('plantilla_detail', kwargs={'pk': self.object.plantilla.id})

# --- Configuración ---


def get_config_activa():
    config, _ = ConfiguracionAlgoritmo.objects.get_or_create(activa=True)
    return config

class ConfiguracionDashboardView(SuperUserRequiredMixin, FormView):
    template_name = 'rostering/config_dashboard.html'
    form_class = ConfiguracionSimpleForm
    success_url = reverse_lazy('config_dashboard')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['config'] = get_config_activa()
        return ctx

    def form_valid(self, form):
        form.save(get_config_activa())
        # Mapear el valor del modo al nombre amigable
        modo_nombres = {
            'RAPIDA': 'Búsqueda Rápida',
            'EQUILIBRADA': 'Búsqueda Equilibrada',
            'PROFUNDA': 'Búsqueda Profunda',
            'PERSONALIZADA': 'Búsqueda Personalizada',
        }
        modo_nombre = modo_nombres.get(form.cleaned_data['modo'], form.cleaned_data['modo'])
        messages.success(self.request, f"Configuración actualizada a {modo_nombre}")
        return super().form_valid(form)

class ConfiguracionSimpleView(SuperUserRequiredMixin, FormView):
    template_name = 'rostering/config_simple.html'
    form_class = ConfiguracionSimpleForm
    success_url = reverse_lazy('config_dashboard')

    def form_valid(self, form):
        form.save(get_config_activa())
        # Mapear el valor del modo al nombre amigable
        modo_nombres = {
            'RAPIDA': 'Búsqueda Rápida',
            'EQUILIBRADA': 'Búsqueda Equilibrada',
            'PROFUNDA': 'Búsqueda Profunda',
            'PERSONALIZADA': 'Búsqueda Personalizada',
        }
        modo_nombre = modo_nombres.get(form.cleaned_data['modo'], form.cleaned_data['modo'])
        messages.success(self.request, f"Configuración actualizada a {modo_nombre}")
        return super().form_valid(form)

class ConfiguracionAvanzadaView(SuperUserRequiredMixin, UpdateView):
    model = ConfiguracionAlgoritmo
    form_class = ConfiguracionAvanzadaForm
    template_name = 'rostering/config_avanzada.html'
    success_url = reverse_lazy('config_dashboard')

    def get_object(self):
        return get_config_activa()

    def form_valid(self, form):
        messages.success(self.request, "Parámetros avanzados guardados.")
        return super().form_valid(form)
        

# Imports necesarios para PDF
from django.template.loader import render_to_string
from django.http import HttpResponse
from django.db.models import Sum
from datetime import timedelta
from weasyprint import HTML

def exportar_cronograma_pdf(request, cronograma_id):
    cronograma = get_object_or_404(Cronograma, pk=cronograma_id)
    
    # --- BLOQUEO DE SEGURIDAD ---
    if cronograma.estado == 'FALLIDO':
        messages.error(request, "No se puede exportar un cronograma fallido.")
        return redirect('cronograma_analisis', pk=cronograma.id)
    # ----------------------------
    
    # 1. Generar encabezado de días en ESPAÑOL manual
    # ... (El resto de tu código de PDF sigue igual) ...
    # ...
    # 1. Generar encabezado de días en ESPAÑOL manual
    # Python: 0=Lunes, 6=Domingo
    letras_dias = ['L', 'M', 'M', 'J', 'V', 'S', 'D']
    
    dias_encabezado = []
    fecha_iter = cronograma.fecha_inicio
    
    while fecha_iter <= cronograma.fecha_fin:
        dias_encabezado.append({
            'fecha': fecha_iter,
            'letra': letras_dias[fecha_iter.weekday()], # Forzamos la letra en español
            'dia_num': fecha_iter.day
        })
        fecha_iter += timedelta(days=1)
    
    num_dias = len(dias_encabezado)

    # 2. Obtener Asignaciones
    asignaciones = Asignacion.objects.filter(
        cronograma=cronograma
    ).select_related('empleado', 'tipo_turno').order_by('empleado__id', 'fecha')

    # 3. Estructurar Matriz (Filas por empleado)
    # Agrupamos asignaciones por empleado
    mapa_empleados = {} # id -> {empleado_obj, celdas: [None]*dias, horas: 0}
    
    # Pre-cargar empleados involucrados (aunque no tengan turno, deberían aparecer si están en la lista original, 
    # pero por simplicidad usamos los que tienen asignación o buscamos los activos de la especialidad)
    # Aquí buscamos todos los activos de la especialidad para que aparezcan aunque no tengan turnos
    empleados_base = Empleado.objects.filter(especialidad=cronograma.especialidad, activo=True).order_by('nombre_completo')

    # Extraemos info de límites del reporte JSON guardado, si existe
    datos_reporte = cronograma.reporte_analisis.get('datos_equidad', {})
    nombres_cortos = datos_reporte.get('nombres_cortos', [])
    nombres_largos = datos_reporte.get('nombres_profesionales', [])
    limites = datos_reporte.get('limites_contractuales', [])
    
    # Mapa auxiliar nombre -> limites
    mapa_limites = {nom: lim for nom, lim in zip(nombres_largos, limites)}

    filas = []
    
    for emp in empleados_base:
        # Recuperar limites guardados o calcular al vuelo
        lims = mapa_limites.get(emp.nombre_completo, [0, 0])
        if lims == [0, 0]: # Fallback si no está en el reporte json
             # Lógica simplificada de fallback
             factor = num_dias / 30
             lims = [emp.min_turnos_mensuales*12*factor, emp.max_turnos_mensuales*12*factor]

        # Nombre corto para la tabla (Generación simple)
        partes = emp.nombre_completo.split()
        n_corto = f"{partes[-1]}, {partes[0][0]}." if len(partes) > 1 else emp.nombre_completo

        # Le inyectamos el atributo temporalmente
        emp.nombre_corto = n_corto
        
        fila = {
            'empleado': emp,
            'celdas': [None] * num_dias,
            'horas_totales': 0,
            'horas_min': lims[0],
            'horas_max': lims[1]
        }
        filas.append(fila)
        mapa_empleados[emp.id] = fila

    # Rellenar celdas
    for asig in asignaciones:
        if asig.empleado.id in mapa_empleados:
            # Calcular índice del día
            delta = (asig.fecha - cronograma.fecha_inicio).days
            if 0 <= delta < num_dias:
                mapa_empleados[asig.empleado.id]['celdas'][delta] = asig
                # Sumar horas (si existe duración, sino 12 por defecto)
                duracion = asig.tipo_turno.duracion_horas if asig.tipo_turno.duracion_horas else 12
                mapa_empleados[asig.empleado.id]['horas_totales'] += float(duracion)

    # 1. Obtener los IDs de tipos de turno que REALMENTE hay en las asignaciones
    # (Esto evita mostrar referencias que no se usaron)
    ids_turnos_usados = asignaciones.values_list('tipo_turno', flat=True).distinct()
    tipos_turno_usados = TipoTurno.objects.filter(id__in=ids_turnos_usados).order_by('nombre')

    # Renderizar
    html_string = render_to_string('rostering/reporte_pdf.html', {
        'cronograma': cronograma,
        'dias_encabezado': dias_encabezado,
        'filas': filas,
        'tipos_turno': tipos_turno_usados,  # <--- AGREGAR ESTO
        'fecha_impresion': timezone.now(),
        'base_url': request.build_absolute_uri('/')
    })
    
    html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))
    pdf_file = html.write_pdf()

    response = HttpResponse(pdf_file, content_type='application/pdf')
    filename = f"Cronograma_{cronograma.id}_{cronograma.fecha_inicio.strftime('%Y%m')}.pdf"
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response

import openpyxl
from datetime import timedelta 
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
# Asegurate de tener tus modelos importados
from .models import Cronograma, Asignacion, Empleado, TipoTurno

def exportar_cronograma_excel(request, cronograma_id):
    cronograma = get_object_or_404(Cronograma, pk=cronograma_id)
    
    # --- BLOQUEO DE SEGURIDAD ---
    if cronograma.estado == 'FALLIDO':
        messages.error(request, "No se puede exportar un cronograma fallido.")
        return redirect('cronograma_analisis', pk=cronograma.id)
    # ----------------------------

    # 1. Crear el libro y la hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Cronograma {cronograma.id}"

    # --- ESTILOS ---
    font_bold = Font(bold=True)
    align_center = Alignment(horizontal='center', vertical='center')
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Colores de fondo (Fills)
    fill_manana = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid") # Amarillo
    fill_tarde = PatternFill(start_color="FD7E14", end_color="FD7E14", fill_type="solid")  # Naranja
    fill_noche = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")  # Azul
    fill_guardia = PatternFill(start_color="198754", end_color="198754", fill_type="solid") # Verde
    fill_enfermeria = PatternFill(start_color="6F42C1", end_color="6F42C1", fill_type="solid") # Violeta
    fill_franco = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid") # Gris claro
    
    # 2. Encabezado Principal
    # CORRECCIÓN: Generamos el nombre dinámicamente con los datos que SÍ existen
    titulo_plan = f"Plan {cronograma.get_especialidad_display()}"
    periodo_str = f"Período: {cronograma.fecha_inicio.strftime('%d/%m/%Y')} al {cronograma.fecha_fin.strftime('%d/%m/%Y')}"
    
    ws.merge_cells('A1:E1')
    ws['A1'] = titulo_plan
    ws['A1'].font = Font(size=14, bold=True, color="0D6EFD")
    
    ws.merge_cells('A2:E2')
    ws['A2'] = periodo_str
    
    # 3. Fila de Días (Encabezados de Tabla)
    ws.cell(row=4, column=1, value="Profesional").font = font_bold
    ws.column_dimensions['A'].width = 25 

    dias = []
    fecha_iter = cronograma.fecha_inicio
    col_idx = 2 
    
    letras_dias = ['L', 'M', 'M', 'J', 'V', 'S', 'D']

    while fecha_iter <= cronograma.fecha_fin:
        letra = letras_dias[fecha_iter.weekday()]
        dia_str = f"{letra}\n{fecha_iter.day}"
        
        cell = ws.cell(row=4, column=col_idx, value=dia_str)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.font = font_bold
        cell.border = border_thin
        
        # Color gris si es finde
        if fecha_iter.weekday() >= 5: 
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 5 
        
        dias.append(fecha_iter)
        fecha_iter += timedelta(days=1) # Ahora sí funciona timedelta
        col_idx += 1

    # 4. Cargar Datos
    asignaciones = Asignacion.objects.filter(
        cronograma=cronograma
    ).select_related('empleado', 'tipo_turno')

    mapa_turnos = {}
    for a in asignaciones:
        mapa_turnos[(a.empleado.id, a.fecha)] = a.tipo_turno

    empleados = Empleado.objects.filter(especialidad=cronograma.especialidad, activo=True).order_by('id')

    # 5. Escribir Filas
    row_idx = 5
    for emp in empleados:
        # Nombre del empleado
        cell_name = ws.cell(row=row_idx, column=1, value=emp.nombre_completo)
        cell_name.border = border_thin
        
        # Turnos
        current_col = 2
        for fecha in dias:
            turno = mapa_turnos.get((emp.id, fecha))
            cell = ws.cell(row=row_idx, column=current_col)
            cell.border = border_thin
            cell.alignment = align_center
            
            if turno:
                nombre_t = turno.nombre
                # Lógica visual de colores (Igual que en PDF)
                sigla = nombre_t[0] 
                fill_to_use = fill_guardia 
                font_color = "FFFFFF" 
                
                if "Mañana" in nombre_t:
                    sigla = "M"
                    fill_to_use = fill_manana
                    font_color = "000000"
                elif "Tarde" in nombre_t:
                    sigla = "T"
                    fill_to_use = fill_tarde
                elif "Noche" in nombre_t:
                    sigla = "N"
                    fill_to_use = fill_noche
                elif "Enferme" in nombre_t:
                    sigla = "E"
                    fill_to_use = fill_enfermeria
                
                cell.value = sigla
                cell.fill = fill_to_use
                cell.font = Font(color=font_color, bold=True)
            else:
                cell.value = "-"
                cell.font = Font(color="CCCCCC")

            current_col += 1
        
        row_idx += 1

    # 6. Devolver respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    # CAMBIO: Usamos fecha inicio Y fin para que sea único
    f_inicio = cronograma.fecha_inicio.strftime('%Y-%m-%d')
    f_fin = cronograma.fecha_fin.strftime('%Y-%m-%d')
    
    filename = f"Cronograma_{f_inicio}_al_{f_fin}.xlsx"
    
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

from django.shortcuts import get_object_or_404, redirect
from django.db import transaction
from django.contrib import messages
from django.contrib.auth.decorators import login_required

@login_required
def duplicar_plantilla(request, pk):
    # 1. Recuperamos la original
    original = get_object_or_404(PlantillaDemanda, pk=pk)
    
    try:
        with transaction.atomic():
            # 2. Generar un nombre único para evitar error de Integrity (Unique Constraint)
            # Intentamos "Copia de [Nombre]", si existe, agregamos un contador.
            nuevo_nombre = f"Copia de {original.nombre}"
            contador = 1
            while PlantillaDemanda.objects.filter(nombre=nuevo_nombre).exists():
                nuevo_nombre = f"Copia de {original.nombre} ({contador})"
                contador += 1
            
            # 3. Crear la nueva plantilla (Copia Superficial)
            nueva_plantilla = PlantillaDemanda.objects.create(
                nombre=nuevo_nombre,
                especialidad=original.especialidad,
                descripcion=original.descripcion
            )

            # 4. Copiar Reglas Semanales (Copia Profunda)
            # Recuperamos las reglas de la original y creamos nuevas apuntando a la nueva_plantilla
            reglas_a_crear = []
            for regla in original.reglas.all():
                reglas_a_crear.append(ReglaDemandaSemanal(
                    plantilla=nueva_plantilla,
                    dias=regla.dias,  # Copiamos la lista de días
                    turno=regla.turno,
                    cantidad_senior=regla.cantidad_senior,
                    cantidad_junior=regla.cantidad_junior
                ))
            # Bulk create es más eficiente que guardar una por una
            ReglaDemandaSemanal.objects.bulk_create(reglas_a_crear)

            # 5. Copiar Excepciones (Copia Profunda)
            excepciones_a_crear = []
            for exc in original.excepciones.all():
                excepciones_a_crear.append(ExcepcionDemanda(
                    plantilla=nueva_plantilla,
                    fecha=exc.fecha,
                    turno=exc.turno,
                    cantidad_senior=exc.cantidad_senior,
                    cantidad_junior=exc.cantidad_junior,
                    motivo=exc.motivo
                ))
            ExcepcionDemanda.objects.bulk_create(excepciones_a_crear)

        messages.success(request, f"Plantilla duplicada exitosamente como '{nuevo_nombre}'.")
        
        # 6. Redirigimos directamente a EDITAR la nueva copia, por si quiere cambiarle el nombre
        return redirect('plantilla_update', pk=nueva_plantilla.pk)

    except Exception as e:
        messages.error(request, f"Error al duplicar la plantilla: {str(e)}")
        return redirect('plantilla_list')


# ==============================================================================
# VISTAS AJAX PARA REGLAS DE DEMANDA (EDICIÓN INLINE)
# ==============================================================================

@login_required
@require_POST
@csrf_exempt
def api_crear_regla(request, plantilla_id):
    """AJAX: Crear una nueva regla de demanda."""
    try:
        plantilla = PlantillaDemanda.objects.get(pk=plantilla_id)
        
        turno_id = request.POST.get('turno_id')
        cantidad_senior = int(request.POST.get('cantidad_senior', 0))
        cantidad_junior = int(request.POST.get('cantidad_junior', 0))
        es_excepcion = request.POST.get('es_excepcion', '0') == '1'
        dias = request.POST.getlist('dias[]')  # Lista de días seleccionados
        dias = [int(d) for d in dias if d]
        
        if not dias:
            return JsonResponse({'error': 'Debe seleccionar al menos un día'}, status=400)
        
        turno = TipoTurno.objects.get(pk=turno_id)
        
        # Validar consistencia de especialidad
        if turno.especialidad != plantilla.especialidad:
            return JsonResponse({'error': 'El turno no corresponde a la especialidad'}, status=400)
        
        # Resolver conflictos
        otras_reglas = ReglaDemandaSemanal.objects.filter(
            plantilla=plantilla,
            turno=turno
        )
        
        for regla in otras_reglas:
            dias_actuales = regla.dias or []
            dias_actualizados = [d for d in dias_actuales if d not in dias]
            
            if dias_actualizados != dias_actuales:
                if dias_actualizados:
                    regla.dias = dias_actualizados
                    regla.save()
                else:
                    regla.delete()
        
        # Crear la nueva regla
        nueva_regla = ReglaDemandaSemanal.objects.create(
            plantilla=plantilla,
            turno=turno,
            dias=dias,
            cantidad_senior=cantidad_senior,
            cantidad_junior=cantidad_junior,
            es_excepcion=es_excepcion
        )
        
        # Retornar los datos en JSON para actualizar la tabla
        return JsonResponse({
            'success': True,
            'regla': {
                'id': nueva_regla.pk,
                'turno_nombre': turno.nombre,
                'turno_abreviatura': turno.abreviatura,
                'cantidad_senior': nueva_regla.cantidad_senior,
                'cantidad_junior': nueva_regla.cantidad_junior,
                'dias': nueva_regla.dias,
                'dias_display': nueva_regla.get_dias_display(),
                'es_excepcion': nueva_regla.es_excepcion
            }
        })
    
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@require_POST
@csrf_exempt
def api_actualizar_regla(request, regla_id):
    """AJAX: Actualizar una regla existente."""
    try:
        regla = ReglaDemandaSemanal.objects.get(pk=regla_id)
        
        cantidad_senior = int(request.POST.get('cantidad_senior', regla.cantidad_senior))
        cantidad_junior = int(request.POST.get('cantidad_junior', regla.cantidad_junior))
        es_excepcion = request.POST.get('es_excepcion', '0') == '1'
        dias = request.POST.getlist('dias[]')
        dias = [int(d) for d in dias if d]
        
        if not dias:
            return JsonResponse({'error': 'Debe seleccionar al menos un día'}, status=400)
        
        # Resolver conflictos
        otras_reglas = ReglaDemandaSemanal.objects.filter(
            plantilla=regla.plantilla,
            turno=regla.turno
        ).exclude(pk=regla_id)
        
        for otra in otras_reglas:
            dias_actuales = otra.dias or []
            dias_actualizados = [d for d in dias_actuales if d not in dias]
            
            if dias_actualizados != dias_actuales:
                if dias_actualizados:
                    otra.dias = dias_actualizados
                    otra.save()
                else:
                    otra.delete()
        
        # Actualizar la regla
        regla.cantidad_senior = cantidad_senior
        regla.cantidad_junior = cantidad_junior
        regla.dias = dias
        regla.es_excepcion = es_excepcion
        regla.save()
        
        return JsonResponse({
            'success': True,
            'regla': {
                'cantidad_senior': regla.cantidad_senior,
                'cantidad_junior': regla.cantidad_junior,
                'dias': regla.dias,
                'dias_display': regla.get_dias_display(),
                'es_excepcion': regla.es_excepcion
            }
        })
    
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@require_POST
@csrf_exempt
def api_eliminar_regla(request, regla_id):
    """AJAX: Eliminar una regla."""
    try:
        regla = ReglaDemandaSemanal.objects.get(pk=regla_id)
        regla.delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)